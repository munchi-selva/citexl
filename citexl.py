#!/usr/bin/env python3
# -*- coding: utf-8 -*-

################################################################################
# Library of functions for manipulating a book citations spreadsheet.
################################################################################

################################################################################
# Imports
################################################################################
from __future__ import unicode_literals
# Ensures Unicode string compatibility # in Python 2/3
import argparse
import itertools
import json
import re
import sys
from collections import Counter
from collections import namedtuple
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.cell.cell import Cell

from os import path
from enum import IntEnum    # Backported to python 2.7 by https://pypi.org/project/enum34

from pprint import pprint   # Pretty printing of lists, tuples, etc.

import sys
if sys.version_info.major == 2:
    from cjklib import characterlookup

import ccdict

#
# Temporary hack, a copy of .pythonrc config to enable autocomplete when
# running interactively
#
try:
    import readline
except ImportError:
    print("Module readline not available.")
else:
    import rlcompleter
    readline.parse_and_bind("tab: complete")

###############################################################################

###############################################################################
# Constants
###############################################################################

######################################
# Default source spreadsheet file name
######################################
SOURCE_FILE     = "/mnt/d/Books_and_Literature/Notes/Cha/Duke_of_Mount_Deer.xlsx"

#######################
# Worksheet style names
#######################
STYLE_GENERAL   = 'BookRef_General'
STYLE_LINK      = 'BookRef_Link'

#########################################################################
# Citation fields names:
# values are retrieved directly or generated from a citation worksheet
#########################################################################
CITE_FLD_CHAPTER        = "回"                  # Generated
CITE_FLD_PAGE           = "頁"                  # Retrieved/generated
CITE_FLD_LINE           = "直行"                # Retrieved/generated
CITE_FLD_INSTANCE       = "instance_number"     # Generated
CITE_FLD_CITE_TEXT      = "引句"                # Retrieved
CITE_FLD_CATEGORY       = "範疇"                # Retrieved
CITE_FLD_TOPIC          = "題"                  # Retrieved
CITE_FLD_PHRASE         = "字詞"                # Retrieved
CITE_FLD_JYUTPING       = "粵拼"                # Retrieved
CITE_FLD_DEFN           = "定義"                # Retrieved/generated
CITE_FLD_ID             = "cite_id"             # Generated
CITE_FLD_LBL_SHORT      = "cit_label_short"     # Generated
CITE_FLD_LBL_VERBOSE    = "cit_label_verbose"   # Generated
CITE_FLD_COUNT          = "cit_count"           # Generated
CITE_FLD_PHRASE_FIRST_INSTANCE  = "cit_phrase_first_instance"   # Generated

#
# Get a list of all citation fields
#
CITE_FLDS = [eval(fld) for fld in list(locals().keys()) if  re.match("CITE_FLD_.*", fld)]


###############################################
# Mapping between citation and CantoDict fields
###############################################
CiteFldToDictFld = {
    CITE_FLD_PHRASE:     ccdict.DE_FLD_TRAD,
    CITE_FLD_JYUTPING:   ccdict.DE_FLD_JYUTPING,
    CITE_FLD_DEFN:       ccdict.DE_FLD_ENGLISH
}


##########################################################################
# Named tuple definition relevant to displaying values of a citation field
##########################################################################
CiteFldDefn = namedtuple("CiteFldDefn", "name disp_width")


###############################################################################
# Citation types enum
###############################################################################
CitType = IntEnum('CitType',  'CT_NONE        \
                               CT_ALL         \
                               CT_DEFN        \
                               CT_REFERRING   \
                               CT_COUNT',

                               start = -1)
###############################################################################

##############################
# Names of citation worksheets
##############################
CITATION_SHEETS = [
    '新序',
    '一圖',
    '一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
    '二圖',
    '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十',
    '三圖',
    '二十一', '二十二', '二十三', '二十四', '二十五', '二十六', '二十七', '二十八', '二十九', '三十',
    '四圖',
    '三十一', '三十二', '三十三', '三十四', '三十五', '三十六', '三十七', '三十八', '三十九', '四十',
    '五圖',
    '四十一', '四十二', '四十三', '四十四', '四十五', '四十六', '四十七', '四十八', '四十九', '五十',
    '附錄', '後記'
]

########################################################################
# Component separators for defined name identifiers and reference labels
########################################################################
DEF_NAME_ID_SEP = '_'
REF_LABEL_SEP   = ';'
###############################################################################

#################################################
# Chinese character shape decomposition constants
#################################################
CJK_SHAPE_LTR   = u'\u2ff0'         # ⿰    Left to right
CJK_SHAPE_ATB   = u'\u2ff1'         # ⿱    Above to below
CJK_SHAPE_LMR   = u'\u2ff2'         # ⿲    Left to middle and right
CJK_SHAPE_AMB   = u'\u2ff3'         # ⿳    Above to middle and below
CJK_SHAPE_SURR  = u'\u2ff4'         # ⿴    Full surrond
CJK_SHAPE_SA    = u'\u2ff5'         # ⿵    Surround from above
CJK_SHAPE_SB    = u'\u2ff6'         # ⿶    Surround from below
CJK_SHAPE_SL    = u'\u2ff7'         # ⿷    Surround from left
CJK_SHAPE_SUL   = u'\u2ff8'         # ⿸    Surround from upper left
CJK_SHAPE_SUR   = u'\u2ff9'         # ⿹    Surround from upper right
CJK_SHAPE_SLL   = u'\u2ffa'         # ⿺    Surround from lower left
CJK_SHAPE_OL    = u'\u2ffb'         # ⿻    Overlaid

#################################################
# Field names for JSON search files
#################################################
MSEARCH_NAME        = "search_name"
MSEARCH_TERMS       = "search_terms"
MSEARCH_TERM_VALUE  = "search_value"
MSEARCH_TERM_FIELD  = "search_field"
MSEARCH_TERM_USE_RE = "use_re_search"


#############################
# Cantonese dictionary object
#############################
canto_dict =  ccdict.CantoDict("cite_dict.db")


###############################################################################
# Miscellaneous spreadsheet helper functions
###############################################################################

###############################################################################
def find_closest_value(ws,
                       col_letter,
                       row_number):
    # type: (Worksheet, str, int) -> (str, int)
    """
    Finds the first non-empty value that appears in the given column, at or
    above the specified row and the row's rank among those sharing that value,
    e.g. if row_number == 5, and the nearest non-empty value is in row 2, rank = 4

    :param  ws:         The worksheet
    :param  col_letter: Column letter
    :param  row_number: Row number (1-based)
    :returns: The value for a given column and row, and the row's rank among
              those sharing a value for that column.
    """

    # Identify non-empty cells at or above the specified row
    non_empty_cells = [c for c in ws[col_letter][:row_number] if not c.value is None]

    if len(non_empty_cells) != 0:
        return non_empty_cells[-1].value, (row_number - non_empty_cells[-1].row + 1)

    return None, None
###############################################################################


###############################################################################
def header_row(ws):
   # type (Worksheet) -> Tuple
   """
   Returns the header row of a worksheet, assumed to contain column names.

   :param ws:  A worksheet
   :returns:   The header row.
   """
   return ws[1]
###############################################################################


###############################################################################
def get_named_row(ws,
                  row_number):
    # type: (Worksheet, int) -> Dict
    """
    Retrieves a worksheet row, formatted as a column/field name to Cell mapping.

    :param  ws:         A worksheet
    :param  row_number: A 1-based row number
    :returns a mapping between the column names and Cells of the specified row
    """
    return dict(zip([header_cell.value for header_cell in header_row(ws)],
                    ws[row_number]))
###############################################################################


###############################################################################
def style_cell(cell):
    # type: (Cell) -> None
    """
    Assigns the appropriate style to a citation worksheet cell

    :param  cell    The cell to be styled
    :returns: Nothing
    """
    cell.style = STYLE_GENERAL if cell.hyperlink is None else STYLE_LINK
###############################################################################


###############################################################################
def style_cit_sheet(ws):
    # type: (Cell) -> None
    """
    Assigns the appropriate style to all cells in a citation worksheet

    :param  ws  The worksheet
    :returns: Nothing
    """
    for row in ws.iter_rows():
        for cell in row:
            style_cell(cell)
###############################################################################


###############################################################################
# A class that wraps up functionality for managing a citation spreadsheet
###############################################################################
class CitationWB(object):
    ###########################################################################
    def __init__(self,
                 src_file           = SOURCE_FILE,
                 cit_sheet_names    = CITATION_SHEETS,
                 cit_id_fields      = [CiteFldDefn(CITE_FLD_PAGE, 2),
                                       CiteFldDefn(CITE_FLD_LINE, 3)]):
        # type: (str, List) -> CitationWB
        """
        Citation workbook constructor

        :param  src_file:           Path to the source spreadsheet file
        :param  cit_sheet_names:    Potential citation worksheet names
                                    (some may not yet be created/filled in)
        :param  cit_id_fields:      The fields that provide an identifier for
                                    a citation
        """
        self.src_file = src_file
        self.cit_sheet_names = cit_sheet_names
        self.cit_id_fields = cit_id_fields
        self.wb = None
        self.ws_col_dicts = None
        self.wb_links = None
        self.reload()
    ###########################################################################


    ###########################################################################
    def reload(self):
        """
        (Re)loads the citation workbook
        """
        if self.wb:
            self.wb.close()
            self.ws_col_dicts = None
            self.wb_links = None
        self.wb = load_workbook(self.src_file)
    ###########################################################################


    ###########################################################################
    def save_changes(self, filename = None):
        # type: (str) -> None
        """
        Saves the workbook to the chosen file
        :param  filename    The destination filename.
                            If this is not provided, a modified version of the
                            source filename is used.
        :returns Nothing
        """
        if not filename:
            filename = re.sub("([.].*$)", ".mod\g<1>", self.src_file)
        self.wb.save(filename)
    ###########################################################################


    ###########################################################################
    def get_valid_cit_sheet_names(self):
        # type: () -> List[Str]
        """
        Retrieves the names of current (existent!) citation worksheets

        :returns a list of the worksheet names
        """
        return [ws_name for ws_name in self.cit_sheet_names if ws_name in self.wb.sheetnames]
    ###########################################################################


    ###########################################################################
    def get_cit_sheets(self):
        # type: () -> List[Worksheet]
        """
        Retrieves the workbook's citation worksheets

        :returns a list of the citation worksheets
        """
        return [self.wb.get_sheet_by_name(ws_name) for ws_name in self.get_valid_cit_sheet_names()]
    ###########################################################################


    ###########################################################################
    def get_column_mappings(self, ws):
        # type (Worksheet) -> Dict
        """
        Returns the mappings between a worksheet's column names and letters.
        A mappings dictionary is initialised/populated as required.

        :param ws:  The worksheet
        :returns:   A dictionary mapping column names to column letters
        """
        if not self.ws_col_dicts:
            self.ws_col_dicts = dict()

        if not ws.title in self.ws_col_dicts:
            self.ws_col_dicts[ws.title] = {header_cell.value: header_cell.column_letter for header_cell in header_row(ws)}
        return self.ws_col_dicts[ws.title]
    ###########################################################################


    ###########################################################################
    def get_col_id(self, ws, col_name):
        # type (Worksheet, str) -> str
        """
        Returns the letter corresponding to the column with the given name in
        the specified worksheet.

        :param ws:          The worksheet
        :param col_name:    The column name
        :returns the column's letter
        """
        return self.get_column_mappings(ws)[col_name]
    ###########################################################################


    ###########################################################################
    def get_defn_links(self, ws):
        # type (Worksheet) -> List[str]
        """
        Finds links (potentially defined names in other worksheets) in citation
        worksheet definitions

        :param ws:  A citation worksheet
        :returns the list of link targets in the worksheet
        """
        links = [defn_cell.hyperlink.location for defn_cell
                    in ws[self.get_col_id(ws, CITE_FLD_DEFN)]
                    if defn_cell.hyperlink and defn_cell.hyperlink.location]
        return links
    ###########################################################################


    ###########################################################################
    def get_link_counts(self):
        # type () -> Counter
        """
        :returns a Counter, mapping links to number of occurrences
        """

        if not self.wb_links:
            #
            # Build the list of links across all citation worksheets
            #
            self.wb_links = list()
            for ws in self.get_cit_sheets():
                self.wb_links.extend(self.get_defn_links(ws))

        #
        # Return the mapping between links and occurrences
        #
        return Counter(self.wb_links)
    ###########################################################################


    ###########################################################################
    def style_workbook(self):
        # type: () -> None
        """
        Assigns the appropriate style to all the workbook's citation worksheets
        :param  wb  The workbook
        :returns: Nothing
        """
        for ws in self.get_cit_sheets():
            style_cit_sheet(ws)
    ###########################################################################


    ###########################################################################
    def get_cit_type(self,
                     cit_row):
        # type: (Dict) -> CitType
        """
        Identifies the a citation row's type

        :param  cit_row:    The citation row
        :returns the citation's type
        """
        cit_type = CitType.CT_NONE
        defn_cell = cit_row[CITE_FLD_DEFN] if cit_row[CITE_FLD_DEFN] else None
        if defn_cell:
            if defn_cell.hyperlink and \
               defn_cell.hyperlink.location in self.wb.defined_names:
                cit_type = CitType.CT_REFERRING
            else:
                cit_type = CitType.CT_DEFN
        return cit_type
    ###########################################################################


    ###########################################################################
    def get_cit_row_values(self,
                           ws,
                           row_number,
                           fields_to_fill = [CITE_FLD_CHAPTER, \
                                             CITE_FLD_PAGE, \
                                             CITE_FLD_LINE, \
                                             CITE_FLD_INSTANCE, \
                                             CITE_FLD_ID, \
                                             CITE_FLD_LBL_SHORT, \
                                             CITE_FLD_LBL_VERBOSE]):
        # type: (Worksheet, int, List[str]) -> Dict
        """
        Retrieves a citation row, formatted as field name to value mappings

        :param  ws:             A citation worksheet
        :param  row_number:     A 1-based row number
        :param  fields_to_fill: Fields that should be generated/filled in if the
                                worksheet does not provide a (direct) value
        :returns a field name to value mapping for the relevant citation row
        """
        return self.get_cit_values(get_named_row(ws, row_number), fields_to_fill)
    ###########################################################################


    ###########################################################################
    def get_cit_values(self,
                       cit_row,
                       fields_to_fill = [CITE_FLD_CHAPTER, \
                                         CITE_FLD_PAGE, \
                                         CITE_FLD_LINE, \
                                         CITE_FLD_INSTANCE, \
                                         CITE_FLD_ID, \
                                         CITE_FLD_LBL_SHORT, \
                                         CITE_FLD_LBL_VERBOSE]):
        # type: (Dict, int) -> Dict
        """
        Retrieves a citation row's values, formatted as field name to value mappings

        :param  cit_row:        The citation row
        :param  fields_to_fill: Fields to be generated/filled in if the
                                worksheet does not provide a (direct) value
        :returns a field name to value mapping for the specified row.
        """
        ws          = cit_row[CITE_FLD_PHRASE].parent
        row_number  = cit_row[CITE_FLD_PHRASE].row
        cit_values  = dict([(cit_data[0], cit_data[1].value) for cit_data in cit_row.items()])

        cit_sheet_name = ws.title
        # The worksheet name corresponds to the chapter numbers, potentially
        # prefixed with part/volume numbers.
        # In this case the components of the chapter number are separated by
        # DEF_NAME_ID_SEP.

        #
        # Identify the components of the citation row's ID, and how these
        # appear in the row's labels
        #
        cit_id_comps    = list()
        cit_label_strs  = [cit_sheet_name.replace(DEF_NAME_ID_SEP, REF_LABEL_SEP)]
        cit_id_strs     = [cit_sheet_name]
        for id_field in self.cit_id_fields:
            fld_name        = id_field.name
            fld_format_str  = "{:0" + str(id_field.disp_width) + "d}"

            id_comp_val, id_comp_instance = find_closest_value(ws, self.get_col_id(ws, fld_name), row_number)

            cit_id_comps.append((id_comp_val, fld_format_str))
            cit_label_strs.append(str(id_comp_val))
            cit_id_strs.append(fld_format_str.format(id_comp_val))

            if fld_name in fields_to_fill:
                #
                # Make the field value available in the name/value mapping
                #
                cit_values[fld_name] = id_comp_val

        #
        # The instance number for the final ID component gives the instance
        # number for the entire citation row
        #
        cit_id_strs.append(str(id_comp_instance))

        cit_label_short = REF_LABEL_SEP.join(cit_label_strs)
        cit_id = DEF_NAME_ID_SEP.join(cit_id_strs)

        cit_values[CITE_FLD_CHAPTER] = ws.title
        if CITE_FLD_INSTANCE in fields_to_fill:
            cit_values[CITE_FLD_INSTANCE] = cit_id_comps[-1][1]
        if CITE_FLD_CITE_TEXT in fields_to_fill:
            cit_values[CITE_FLD_CITE_TEXT], _ = find_closest_value(ws, self.get_col_id(ws, CITE_FLD_CITE_TEXT), row_number)
        if CITE_FLD_ID in fields_to_fill:
            cit_values[CITE_FLD_ID] = cit_id
        if CITE_FLD_LBL_SHORT in fields_to_fill:
            cit_values[CITE_FLD_LBL_SHORT] = cit_label_short
        if CITE_FLD_LBL_VERBOSE in fields_to_fill:
            cit_values[CITE_FLD_LBL_VERBOSE] = "{}!{}".format(cit_label_short, row_number)
        if CITE_FLD_COUNT in fields_to_fill:
            link_counter = self.get_link_counts()
            if self.get_cit_type(cit_row) == CitType.CT_REFERRING:
                #
                # For citations that refer to another (source) citation,
                # retrieve the source citation's counter
                #
                cit_values[CITE_FLD_COUNT] = link_counter[cit_row[CITE_FLD_DEFN].hyperlink.location] + 1
            elif self.get_cit_type(cit_row) == CitType.CT_DEFN:
                cit_values[CITE_FLD_COUNT] = link_counter[cit_id] + 1
        if CITE_FLD_DEFN in fields_to_fill:
            if self.get_cit_type(cit_row) == CitType.CT_REFERRING:
                #
                # For citations that refer to another (source) citation,
                # retrieve the definition and Jyutping value from the
                # source citation
                #
                defn_src_id = cit_row[CITE_FLD_DEFN].hyperlink.location

                #
                # Extract source citation location from its defined name data
                # Example destination format: ("十一", "$F$99")
                #
                defn_src_locator    = list(self.wb.defined_names[defn_src_id].destinations)[0]
                defn_src_ws         = self.wb.get_sheet_by_name(defn_src_locator[0])
                defn_src_row_number = defn_src_locator[1].split("$")[2]

                defn_src_row = get_named_row(defn_src_ws, defn_src_row_number)

                cit_values[CITE_FLD_DEFN] = defn_src_row[CITE_FLD_DEFN].value
                cit_values[CITE_FLD_JYUTPING] = defn_src_row[CITE_FLD_JYUTPING].value
        if CITE_FLD_PHRASE_FIRST_INSTANCE in fields_to_fill:
            if self.get_cit_type(cit_row) == CitType.CT_REFERRING:
                cit_values[CITE_FLD_PHRASE_FIRST_INSTANCE] = cit_row[CITE_FLD_DEFN].hyperlink.location
        return cit_values
    ###########################################################################


    ###########################################################################
    @staticmethod
    def format_cit_value(cit_value,
                         fld_name):
        # type: (str, str) -> (str, str)
        """
        Generates a formatted version of a citation field value, plus its trailing
        delimiter

        :param  cit_value:  Raw citation field value
        :param  fld_name:   Name of the citation field
        :returns the formatted field value and trailing delimiter.
        """
        display_value = ""
        display_delim = ""
        if fld_name == CITE_FLD_CITE_TEXT:
            display_value = "\"{}\"".format(cit_value) if cit_value else "-"
            display_delim = ' '
        elif fld_name == CITE_FLD_CATEGORY:
            display_value = "<{}>".format(cit_value if cit_value else "-")
            display_delim = " "
        elif fld_name == CITE_FLD_TOPIC:
            display_value = "[{}]".format(cit_value) if cit_value else ""
            display_delim = " "
        elif fld_name == CITE_FLD_PHRASE or fld_name == CITE_FLD_DEFN:
            display_value = cit_value if cit_value else ""
            display_delim = "\t"
        elif fld_name == CITE_FLD_JYUTPING:
            display_value = "({})".format(cit_value) if cit_value else ""
            display_delim = "\t"
        elif fld_name == CITE_FLD_COUNT:
            display_value = "({})".format(cit_value) if cit_value is not None else ""
            display_delim = " "
        elif fld_name == CITE_FLD_LBL_VERBOSE:
            display_value = cit_value if cit_value else ""
            display_delim = "\t"
        else:
            display_value = cit_value if cit_value else ""
            display_delim = " "
        return display_value, display_delim
    ###########################################################################


    ###########################################################################
    @staticmethod
    def format_cit_values(cit_values,
                          cit_fields = [CITE_FLD_CATEGORY, CITE_FLD_PHRASE,
                                        CITE_FLD_JYUTPING, CITE_FLD_DEFN]):
        # type: (Dict, List) -> str
        """
        Retrieves a formatted string corresponding to a citation

        :param  cit_values: Field to value mapping for the citation
        :param  cit_fields: Fields to include
        :returns the selected citation values as formatted string
        """
        cit_str = ""
        for cit_field in cit_fields:
            fld_display_value, delim = CitationWB.format_cit_value(cit_values.get(cit_field, None),
                                                                   cit_field)
            cit_str += "{}{}".format(fld_display_value, delim)
        return cit_str
    ###########################################################################


    ###########################################################################
    def format_cit_row(self,
                       cit_row,
                       cit_fields = [CITE_FLD_LBL_VERBOSE, CITE_FLD_CATEGORY,
                                     CITE_FLD_PHRASE,
                                     CITE_FLD_JYUTPING, CITE_FLD_DEFN]):
        # type: (Dict, List) -> str
        """
        Retrieves a formatted string corresponding to a citation

        :param  cit_row:    A citation row
        :param  cit_fields: Fields to include
        :returns the citation as a formatted string
        """

        #
        # Retrieve the mapping between citation field names and values
        #
        cit_values = self.get_cit_values(cit_row, fields_to_fill = cit_fields)
        return CitationWB.format_cit_values(cit_values, cit_fields)
    ###########################################################################


    ###########################################################################
    def build_reference(self,
                        referenced_row,
                        referring_row,
                        overwrite       = False,
                        audit_only      = False):
        # type: (Dict, Dict, bool, bool) -> None
        """
        Builds a reference between two citation rows.
        The referencing is achieved by creating a defined name that includes
        the referenced row's phrase cell, and linking to this name from the
        referring row's phrase cell.

        :param  referenced_row  The referenced row
        :param  referring_row   The referring row
        :param  overwrite       If True, overwrite any existing content in the
                                referring definition cell (if this isn't already the
                                label for the referenced phrase cell)
        :param  audit_only      If True, show/print the actions for building the
                                reference, but do not modify the data
        :returns: Nothing
        """

        #
        # Tracks actions required in order to build the reference
        #
        audit_log = list()

        #
        # Retrieve referenced and referring cells and worksheets
        #
        referenced_cell = referenced_row[CITE_FLD_PHRASE]
        referenced_ws   = referenced_cell.parent
        referring_cell  = referring_row[CITE_FLD_DEFN]
        referring_ws    = referring_cell.parent

        #
        # Retrieve an identifier for the defined name, and the label for referring
        # to the referenced cell (to be displayed in the referring cell)
        #
        referenced_cit_vals = self.get_cit_values(referenced_row, CITE_FLDS)
        def_name_id = referenced_cit_vals[CITE_FLD_ID]
        label = referenced_cit_vals[CITE_FLD_LBL_SHORT]

        if not def_name_id is None and not label is None:
            workbook = referenced_ws.parent
            if not def_name_id in workbook.defined_names:
                #
                # Create the defined name
                #
                def_name_destination = '{}!${}${}'.format(referenced_ws.title,
                                                          referenced_cell.column_letter,
                                                          referenced_cell.row)
                def_name = DefinedName(name = def_name_id,
                                       attr_text = def_name_destination)
                audit_log.append("Create defined name: {}: {}".format(def_name_destination, def_name_id))

                if not audit_only:
                    workbook.defined_names.append(def_name)

        write_needed = referring_cell.value is None or (overwrite and referring_cell.value != label)
        referring_cell_loc = referring_cell.coordinate
        if write_needed:
            audit_log.append("{}!{} current value = {}".format(referring_ws.title,
                                                               referring_cell_loc,
                                                               referring_cell.value))
            if not audit_only:
                referring_cell.value = label
                referring_cell.hyperlink = Hyperlink(ref = referring_cell_loc,
                                                     location = def_name_id)

                # Clear Jyutping cell contents
                jyutping_cell       = referring_row[CITE_FLD_JYUTPING]
                jyutping_cell.value = None

                # Assign styles
                style_cell(referring_cell)
                style_cell(jyutping_cell)

        if len(audit_log) != 0:
            print("{}!{} ({}) --> {}!{}".format(
                  referring_ws.title, referring_cell.coordinate, referring_cell.value,
                  referenced_ws.title, referenced_cell.coordinate))
            for audit_msg in audit_log:
                print("\t{}".format(audit_msg))
    ###########################################################################


    ###########################################################################
    def find_matches_in_sheet(self,
                              ws,
                              search_expr,
                              fld_name,
                              do_re_search  = False,
                              cit_type      = CitType.CT_ALL,
                              max_matches   = -1):
        # type: (Worksheet, str/Set(str), str, bool, CitType, int) -> List[Dict]
        """
        Finds citation rows matching conditions on a given field

        :param  search_expr:    Search term/s to be matched, converted to a set if
                                necessary
        :param  fld_name:       Name of the field to be searched
        :param  do_re_search:   If True, treat search terms as regular expressions
        :param  cit_type:       Type of citations to search for
        :param  max_matches:    Maximum number of matches to return
        :returns a list of column name to Cell mappings for the matching rows.
        """

        #
        # Correct formatting of search_expr
        #
        if isinstance(search_expr, str):
            search_expr = {search_expr}

        #
        # Add None to allow searching for blank column values if appropriate
        #
        if "" in search_expr and not do_re_search:
            search_expr.add(None)

        #
        # Build the list of cells in the worksheet matching the search terms
        #
        search_col  = self.get_col_id(ws, fld_name)
        matching_cells = ws[search_col][1:]
        if do_re_search:
            matching_cells  = [cell for cell in matching_cells if cell.value and any(re.search(term, cell.value) for term in search_expr)]
        else:
            matching_cells  = [cell for cell in matching_cells if cell.value in search_expr]

        #
        # Retrieve the corresponding rows
        #
        matching_rows = [get_named_row(ws, cell.row) for cell in matching_cells]

        #
        # Filter based on cell type
        #
        if cit_type != CitType.CT_ALL:
            matching_rows = [r for r in matching_rows if self.get_cit_type(r) == cit_type]

        #
        # Trim the results list to the maximum instances limit
        #
        if max_matches > 0:
            matching_rows = matching_rows[:max_matches]

        return matching_rows
    ###########################################################################


    ###########################################################################
    def find_citations_with_no_def(self,
                                   ws,
                                   min_num_chars = 1,
                                   max_num_chars = 1):
        # type (Worksheet, int, int) -> List[Dict]
        """
        Finds citation rows with no associated definition

        :param ws:              A citation worksheet
        :param min_num_chars:   Minimum number of characters in the phrase
        :param max_num_chars:   Maximum number of characters in the phrase,
                                if 0 no upper limit is imposed on the phrase length
        :returns the list of citation rows with no definition
        """
        cit_rows = self.find_matches_in_sheet(ws, "", CITE_FLD_DEFN)
        if max_num_chars > 0:
            cit_rows = [row for row in cit_rows
                            if row[CITE_FLD_PHRASE].value and
                               len(row[CITE_FLD_PHRASE].value) >= min_num_chars and
                               len(row[CITE_FLD_PHRASE].value) <= max_num_chars]
        else:
            cit_rows = [row for row in cit_rows
                            if row[CITE_FLD_PHRASE].value and
                               len(row[CITE_FLD_PHRASE].value) >= min_num_chars]
        return cit_rows
    ###########################################################################


    ###########################################################################
    def find_matches(self,
                     search_expr,
                     fld_name,
                     do_re_search   = False,
                     cit_type       = CitType.CT_ALL,
                     max_matches    = -1):
        # type: (Workbook, str/List[str], str, bool, CitType, int) -> List[Dict]
        """
        Finds citation rows matching conditions on a given column.

        :param  search_expr:    Search term/s to be matched
        :param  fld_name:       Name of the field to be searched
        :param  do_re_search:   If True, treat search terms as regular expressions
        :param  cit_type:       Type of citations to search for
        :param  max_matches:    Maximum number of matched citations to return
        :returns a list of column name to Cell mappings for the matching rows.
        """
        wb = self.wb

        #
        # Initialise match list
        #
        matching_rows = list()

        #
        # Perform search over all citation sheets
        #
        cit_sheets = self.get_cit_sheets()
        max_ws_matches = max_matches
        for ws in cit_sheets:
            #
            # Check whether the maximum matches limit has been reached
            #
            if max_ws_matches == 0:
                break

            #
            # Find matches in the latest worksheet
            #
            ws_matches = self.find_matches_in_sheet(ws,
                                                    search_expr,
                                                    fld_name,
                                                    do_re_search,
                                                    cit_type,
                                                    max_ws_matches)
            if len(ws_matches) > 0:
                #
                # Add matches to return list and update the next worksheet's limit
                #
                matching_rows += ws_matches

                if max_matches > 0:
                    max_ws_matches -= len(ws_matches)

        return matching_rows
    ###########################################################################


    ###########################################################################
    def find_matches_for_file(self,
                              search_terms_filename,
                              cit_type                 = CitType.CT_DEFN):
        # type (str, CitType) -> Dict
        """
        Find matches in the workbook for search terms specified in a JSON file.
        For search terms with no hits, revert to a dictionary lookup.

        :param  search_terms_filename:  Name of the file containing the terms
        :param  cit_type:               Type of citations to search for
        :returns a dictionary mapping search groups (provided by the JSON file)
                 to search results
        """

        #
        # Initialise the return value
        #
        file_matches = dict()

        with open(search_terms_filename) as search_terms_file:
            search_groups = json.load(search_terms_file)
            for search_group in search_groups:
                search_name = search_group[MSEARCH_NAME]
                search_terms = search_group[MSEARCH_TERMS]
                for search_term in search_terms:
                    #
                    # Retrieve matches for this search group
                    #
                    group_matches = file_matches.get(search_name, list())

                    search_value = search_term.get(MSEARCH_TERM_VALUE)
                    search_field = search_term.get(MSEARCH_TERM_FIELD, CITE_FLD_PHRASE)
                    use_re_search = search_term.get(MSEARCH_TERM_USE_RE, False)
                    cit_matches = self.find_matches(search_value,
                                                    search_field,
                                                    use_re_search,
                                                    max_matches = 1)
                    if len(cit_matches) == 1:
                        cit_values = self.get_cit_values(cit_matches[0], CITE_FLDS)
                        group_matches.append(cit_values)
                    else:
                        #
                        # Generate an artificial citation mapping via a dictionary
                        # search
                        #
                        dict_matches = canto_dict.search_dict(search_value)
                        for dict_match in dict_matches:
                            cit_values = dict()
                            for cite_fld in CiteFldToDictFld.keys():
                                ccdict_field = CiteFldToDictFld.get(cite_fld)
                                if ccdict_field:
                                    cit_values[cite_fld] = dict_match.get(ccdict_field)
                            cit_values[CITE_FLD_COUNT] = 0
                            group_matches.append(cit_values)
                    file_matches[search_name] = group_matches
            return file_matches
    ###########################################################################


    ###########################################################################
    def find_cits_by_shape_and_value(self,
                                     shape,
                                     value,
                                     pos):
        # type (str, str, int) -> List[Dict]
        """
        Finds citations where the cited phrase fits CJK shape conditions

        :param  shape:  Ideographic shape to match
        :param  value:  Value to match within the shape
        :param  pos:
        :returns the list of matching citation rows
        """
        matches = list()

        cjk = characterlookup.CharacterLookup('T')
        sheets = self.get_cit_sheets()
        for ws in sheets:
            ws_matches = [get_named_row(ws, row_number) for row_number in range(2, ws.max_row + 1)]
            ws_matches = [row for row in ws_matches if self.get_cit_type(row) == CitType.CT_DEFN]
            ws_matches = [row for row in ws_matches if row[CITE_FLD_PHRASE].value and
                            len(cjk.getDecompositionEntries(row[CITE_FLD_PHRASE].value)) > 0]

            if not shape:
                radical_index = cjk.getKangxiRadicalIndex(value)
                ws_matches = [row for row in ws_matches if cjk.getCharacterKangxiRadicalIndex(row[CITE_FLD_PHRASE].value) == radical_index]
            else:
                ws_matches = [row for row in ws_matches if cjk.getDecompositionEntries(row[CITE_FLD_PHRASE].value)[0][0] == shape
                    and cjk.getDecompositionEntries(row[CITE_FLD_PHRASE].value)[0][pos+1][0] == value]


            #decomps = [(cell, cjk.getDecompositionEntries(cell.value)) for cell in
                    #phrase_cells if len(cjk.getDecompositionEntries(cell.value)) > 0 and
                    #cjk.getDecompositionEntries(cell.value)[0][0] == shape and
                    #cjk.getDecompositionEntries(cell.value)[0][pos+1][0] == value
                    #]
            #for decomp in decomps:
                #print(decomp)
            matches.extend(ws_matches)

        return [self.get_cit_values(row) for row in matches]
    ###########################################################################


    ###########################################################################
    def get_sheet_cit_values(self,
                             ws_name,
                             cit_fields = CITE_FLDS):
        # type: (str, List(str)) -> List(Dict)
        """
        Get the citation values for a given worksheet

        :parm   ws_name:    Name of a citation worksheet
        :param  cit_fields: TODO

        :returns the worksheet's citation values as a list
        """
        cits = list()
        if ws_name in self.get_valid_cit_sheet_names():
            ws = self.wb.get_sheet_by_name(ws_name)
            for row_number in range(2, ws.max_row + 1):
                cits.append(self.get_cit_row_values(ws, row_number, cit_fields))
        return cits
    ###########################################################################


    ###########################################################################
    def get_all_cit_values(self,
                           cit_fields = CITE_FLDS):
        # type: (List(str)) -> List(Dict)
        """
        Get all citation values for the workbook

        :param  cit_fields: TODO

        :returns the workbook's citation values as a list
        """
        cits = list()
        for ws_name in self.get_valid_cit_sheet_names():
            cits.extend(self.get_sheet_cit_values(ws_name, cit_fields))
        return cits
    ###########################################################################


    ###########################################################################
    def display_cit_values_list(self,
                                cit_values_list,
                                cit_fields      = CITE_FLDS,
                                prefix          = ""):
        # type: (List(Dict), List(str), str) -> None
        """
        Displays a list of citation row values

        :param cit_values_list: A list of citation rows, formatted as field
                                to value mappings
        :param cit_fields:      Fields to display
        :param prefix:          A string to prefix to each citation line
        """
        for cit_values in cit_values_list:
            print("{}{}".format(prefix, CitationWB.format_cit_values(cit_values, cit_fields)))
    ###########################################################################


    ###########################################################################
    def display_matches(self,
                        search_expr,
                        fld_name        = CITE_FLD_PHRASE,
                        do_re_search    = False,
                        cit_type        = CitType.CT_DEFN,
                        max_matches     = -1,
                        cit_fields      = [CITE_FLD_LBL_VERBOSE, CITE_FLD_CATEGORY,
                                           CITE_FLD_PHRASE, CITE_FLD_JYUTPING,
                                           CITE_FLD_DEFN]):
        # type: (str/Set(str), str, bool, CitType, int, List[str]) -> None
        """
        Displays the matches for one or more search terms in the citations workbook

        :param  search_expr:    Search term/s to be matched
        :param  fld_name:       Name of the field to be searched
        :param  do_re_search:   If True, treat search terms as regular expressions
        :param  cit_type:       Type of citations to search for
        :param  max_matches:    Maximum number of matched cells to return
        :param  cit_fields:     Fields to display
        :returns nothing
        """
        cit_values_list = [self.get_cit_values(cit_row, cit_fields)
                               for cit_row
                               in self.find_matches(search_expr, fld_name, do_re_search, cit_type, max_matches)]
        self.display_cit_values_list(cit_values_list, cit_fields)
    ###########################################################################


    ###############################################################################
    def display_matches_for_file(self,
                                 search_terms_filename,
                                 cit_type               = CitType.CT_DEFN,
                                 cit_fields             = [CITE_FLD_LBL_VERBOSE,
                                                           CITE_FLD_COUNT,
                                                           CITE_FLD_CATEGORY,
                                                           CITE_FLD_PHRASE,
                                                           CITE_FLD_JYUTPING,
                                                           CITE_FLD_DEFN]):
        # type (str, bool, str, CitType, List[str], bool) -> None
        """
        Find and display matches in the workbook for search terms specified in
        a JSON file.
        For search terms with no hits in the workbook, revert to a dictionary
        lookup.

        :param  search_terms_filename:  Name of the file containing the terms
        :param  cit_type:               Type of citations to search for
        :param  cit_fields:             Citation fields to show
        :returns Nothing
        """
        file_matches = self.find_matches_for_file(search_terms_filename, cit_type)
        for search_name, citations in file_matches.items():
            print(search_name)
            self.display_cit_values_list(citations, cit_fields, "\t")
    ###############################################################################


    ############################################################################
    def display_multiply_used_defns(self,
                                    cit_fields = [CITE_FLD_COUNT,
                                                  CITE_FLD_PHRASE,
                                                  CITE_FLD_DEFN]):
        # type () -> None
        """
        Shows the definition of phrases that are recorded multiple times in the
        citation workbook.

        :param  cit_fields: Fields to display
        :returns Nothing
        """

        wb = self.wb
        cit_values_list = list()

        #
        # Traverse the link counter in descending number of occurrences,
        # collecting the citations associated with the links for display
        #
        link_counter = self.get_link_counts()
        for link_name, ref_count in link_counter.most_common():
            defined_name = wb.defined_names.get(link_name)
            ws_name, coordinate = defined_name.attr_text.split('!')
            ws = wb.get_sheet_by_name(ws_name)
            cit_row = get_named_row(ws, ws[coordinate].row)
            cit_values_list.append(self.get_cit_values(cit_row, cit_fields))
        self.display_cit_values_list(cit_values_list, cit_fields)
    ############################################################################


    ###########################################################################
    def get_refs_for_ws_phrases(self,
                                ws_name,
                                overwrite,
                                audit_only):
        # type: (str, bool, bool) -> None
        """
        Fills in the references for the citations in a worksheet.
        This requires building a reference to the first occurrence of each
        phrase cited in the workbook.

        :param  ws_name:    A citation worksheet name
        :param  overwrite:  If True, overwrite any existing content in referring
                            cells if these don't already refer to the referenced
                            cells
        :param  audit_only: If True, show/print the actions for building references
                            without modifying any data
        :returns Nothing
        """
        if ws_name in self.get_valid_cit_sheet_names():
            ws = self.wb.get_sheet_by_name(ws_name)
            #
            # Iterate through all non-empty phrase cells of the chapter worksheet
            #
            for phrase_cell in [c for c in ws[self.get_col_id(ws, CITE_FLD_PHRASE)][1:] if c.value]:
                #
                # Find the first cell to define the phrase
                #
                referenced_rows = self.find_matches(phrase_cell.value, CITE_FLD_PHRASE, False, CitType.CT_DEFN, 1)
                referenced_row  = referenced_rows[0] if len(referenced_rows) > 0 else None
                referenced_cell = referenced_row[CITE_FLD_PHRASE] if referenced_row else None

                if referenced_cell and not referenced_cell == phrase_cell:
                    #
                    # Ensure this cell isn't the one providing the definition!
                    #
                    referring_row = get_named_row(ws, phrase_cell.row)
                    self.build_reference(referenced_row, referring_row, overwrite, audit_only)
    ###########################################################################


    ###########################################################################
    @staticmethod
    def fill_defn(cit_row,
                  overwrite = False,
                  audit_only = False):
        # type: (Dict, bool, bool) -> bool
        """
        Fills in the definition (including Jyutping transcription) of a citation
        row.

        :param  cit_row:    A row from a citation worksheet
        :param  overwrite:  If True, overwrites existing definition information
        :param  audit_only: If True, print the definition data without
                            modifying the citation worksheet
        :returns True if definition data was found
        """

        #
        # Check for the existence of a valid phrase cell before proceeding
        #
        phrase_cell = cit_row[CITE_FLD_PHRASE]
        if not phrase_cell or not phrase_cell.value:
            return False

        #
        # Retrieve the worksheet and row via the phrase_cell
        #
        ws          = phrase_cell.parent
        row_number  = phrase_cell.row

        INTRA_DEF_SEP   = ", "
        INTER_DEF_SEP   = ";\n"

        jsonDecoder = json.JSONDecoder()

        #
        # Each search result bundles up a list of English definitions and
        # Jyutping transcriptions corresponding to the phrase
        #
        dict_search_res = canto_dict.search_dict(phrase_cell.value)
        defn_vals = list()
        jyutping_vals = list()
        for search_res in dict_search_res:
            #
            # Generate English and Jyutping strings
            #
            defn_list = jsonDecoder.decode(search_res[ccdict.DE_FLD_ENGLISH])
            defn_list = list(filter(None, defn_list))
            if len(defn_list) != 0:
                defn_vals.append(INTRA_DEF_SEP.join(defn_list))
            jyutping_list = jsonDecoder.decode(search_res[ccdict.DE_FLD_JYUTPING])
            jyutping_list = list(filter(None, jyutping_list))
            if len(jyutping_list) == 0:
                jyutping_list.append("?")
            jyutping_vals.append(INTRA_DEF_SEP.join(jyutping_list))

        jyut_cell = cit_row[CITE_FLD_JYUTPING]
        defn_cell = cit_row[CITE_FLD_DEFN]

        print("{}:\t{}".format(phrase_cell.row, phrase_cell.value))
        print(INTER_DEF_SEP.join(["\t{}".format(jyutping) for jyutping in jyutping_vals]))
        print(INTER_DEF_SEP.join(["\t{}".format(defn) for defn in defn_vals]))

        if not audit_only:
            if not jyut_cell.value or overwrite:
                jyut_cell.value = INTER_DEF_SEP.join(jyutping_vals)
                style_cell(jyut_cell)
            if not defn_cell.value or overwrite:
                defn_cell.value = INTER_DEF_SEP.join(defn_vals)
                style_cell(defn_cell)

        return len(defn_vals) > 0 or len(jyutping_vals) > 0
    ###########################################################################


    ###########################################################################
    def fill_in_sheet(self,
                      ws_name,
                      overwrite = False,
                      audit_only = False):
        # type: (str, bool) -> None
        """
        Fills in a citation worksheet by:
            1. Building links to prior citations that define the same phrase
            2. Performing a dictionary lookup for first occurrences of phrases
            3. Displaying cited phrases that still require a definitions after 1/2.

        :parm   ws_name:    Name of a citation worksheet
        :param  overwrite:  If True, overwrite any existing content in cells to be
                            filled out
        :returns the list of citations that still have no definition
        """
        cits_with_no_def = list()
        if ws_name in self.get_valid_cit_sheet_names():
            ws = self.wb.get_sheet_by_name(ws_name)
            self.get_refs_for_ws_phrases(ws_name, overwrite, audit_only)

            fields_to_fill = [CITE_FLD_CHAPTER, CITE_FLD_PAGE, CITE_FLD_LINE, \
                              CITE_FLD_INSTANCE, \
                              CITE_FLD_ID, CITE_FLD_LBL_SHORT, CITE_FLD_LBL_VERBOSE,
                              CITE_FLD_CITE_TEXT,
                              CITE_FLD_COUNT]


            #
            # Attempt to fill in definitions/Jyutping for single character phrases
            #
            cits_to_fill = self.find_citations_with_no_def(ws)
            for cit_row in cits_to_fill:
                if not CitationWB.fill_defn(cit_row, overwrite, False):
                    cits_with_no_def.append(self.get_cit_values(cit_row, fields_to_fill = fields_to_fill))

            #
            # Repeat for multi-character phrases
            #
            cits_to_fill = self.find_citations_with_no_def(ws, 2, -1)
            for cit_row in cits_to_fill:
                if not CitationWB.fill_defn(cit_row, overwrite, False):
                    cits_with_no_def.append(self.get_cit_values(cit_row, fields_to_fill = fields_to_fill))

            if len(cits_with_no_def) != 0:
                print("Definition still required...")
                self.display_cit_values_list(cits_with_no_def, [CITE_FLD_LBL_VERBOSE, CITE_FLD_PHRASE, CITE_FLD_CITE_TEXT], "\t")
        return cits_with_no_def
    ###########################################################################


    ###########################################################################
    def fill_in_last_sheet(self,
                           overwrite = False):
        # type: (bool) -> None
        """
        Fills in the latest citation worksheet based on existing definitions,
        then shows the phrases that require definitions.

        :param  overwrite:  If True, overwrite any existing content in cells
                            to be filled out
        :returns the list of citations that still have no definition
        """
        last_sheet_name = self.get_valid_cit_sheet_names()[-1]
        return self.fill_in_sheet(last_sheet_name, overwrite)
###############################################################################



###############################################################################
def show_char_decomposition(c):
    # type (char) -> Nothing
    """
    Shows the CJK shape decomposition of a character

    :param  c:  A character
    """
    cjk = characterlookup.CharacterLookup('T')
    decs = cjk.getDecompositionEntries(c)
    for dec in decs:
        print(dec)

###############################################################################


###############################################################################
def main():
    """
    TODO

    :returns: None
    """

###############################################################################


if __name__ == "__main__":
    main()

    citewb = CitationWB()
#   cits_with_no_defs = citewb.fill_in_last_sheet(True)
#   cits_with_no_defs = list()
#   for sheet_name in citewb.cit_sheet_names:
#       cits_with_no_defs.extend(citewb.fill_in_sheet(sheet_name, True))
#   citewb.save_changes()

#   cits = citewb.get_all_cit_values()
#   cit_file = open("../citeview/Duke_of_Mount_Deer.cit", "w")
#   json.dump(cits, cit_file)

#   citewb.display_matches_for_file("confounds.match",
#                                   cit_fields = CITE_FLDS)

#   citewb.display_matches("囉", cit_fields = [CITE_FLD_LBL_VERBOSE, CITE_FLD_PHRASE, CITE_FLD_CITE_TEXT, CITE_FLD_DEFN], max_matches = 3)
#   citewb.display_matches("囉", cit_fields = [CITE_FLD_LBL_VERBOSE, CITE_FLD_PHRASE, CITE_FLD_CITE_TEXT, CITE_FLD_DEFN], cit_type = CitType.CT_ALL,
#           max_matches = 3)
#   citewb.display_matches("囉", cit_fields = [CITE_FLD_LBL_VERBOSE, CITE_FLD_PHRASE, CITE_FLD_CITE_TEXT, CITE_FLD_DEFN], cit_type = CitType.CT_REFERRING)
#   citewb.display_matches("", CITE_FLD_TOPIC, do_re_search = True, cit_fields = CITE_FLDS)
#   citewb.display_matches("..武", CITE_FLD_TOPIC, do_re_search = True, cit_fields = CITE_FLDS)
#   citewb.display_matches("囉", cit_fields = CITE_FLDS)
#   citewb.display_matches("..武", CITE_FLD_TOPIC, do_re_search = True, cit_fields = [CITE_FLD_LBL_VERBOSE, CITE_FLD_CITE_TEXT, CITE_FLD_CATEGORY, CITE_FLD_TOPIC, CITE_FLD_PHRASE, CITE_FLD_DEFN])
#   citewb.display_matches("..武", CITE_FLD_TOPIC, do_re_search = True, cit_fields = [CITE_FLD_CITE_TEXT, CITE_FLD_LBL_VERBOSE, CITE_FLD_CATEGORY, CITE_FLD_TOPIC, CITE_FLD_PHRASE, CITE_FLD_DEFN])

#   citewb.display_multiply_used_defns()
    citewb.display_matches("情節", cit_fields = [CITE_FLD_ID, CITE_FLD_LBL_SHORT, CITE_FLD_PAGE, CITE_FLD_LINE, CITE_FLD_INSTANCE, CITE_FLD_PHRASE, CITE_FLD_JYUTPING, CITE_FLD_DEFN])

    hist_citewb = CitationWB(src_file = '/mnt/d/Books_and_Literature/Notes/Liu/Understanding_China.xlsx',
                             cit_sheet_names    = ['一', '二', '三', '四'],
                             cit_id_fields      = [CiteFldDefn(CITE_FLD_PAGE, 3),
                                                   CiteFldDefn("面板", 2),
                                                   CiteFldDefn("段", 2)])
    hist_citewb.fill_in_sheet('二', True)
    hist_citewb.save_changes()

    gu_citewb = CitationWB(src_file = '/mnt/d/Books_and_Literature/Notes/Gu/The_Sword_God_Smiles.xlsx',
                           cit_sheet_names = ['序',
                                              '一_一', '一_二', '一_三', '一_四', '一_五', 
                                              '一_六', '一_七', '一_八', '一_九', '一_十', 
                                              '二_一', ' 二_二', '二_三', '二_四', '二_五',
                                              '二_六', '二_七', '二_八', '二_九', '二_十'],
                           cit_id_fields = [CiteFldDefn("段", 2),
                                            CiteFldDefn(CITE_FLD_PAGE, 3),
                                            CiteFldDefn(CITE_FLD_LINE, 2)])
    if  sys.version_info.major ==  2:
        show_char_decomposition('彆')
        matches = citewb.find_cits_by_shape_and_value(CJK_SHAPE_LTR, '口', 0)
        citewb.display_cit_values_list(matches)
